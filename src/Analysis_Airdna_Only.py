from openpyxl import load_workbook
import requests
from openpyxl.comments import Comment
import json

def Airdna(cityId='', regionId=''):
    headers = {
        'authority': 'api.airdna.co',
        'sec-ch-ua': '"Chromium";v="94", "Google Chrome";v="94", ";Not A Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/94.0.4606.71 Safari/537.36',
        'sec-ch-ua-platform': '"Windows"',
        'accept': '*/*',
        'origin': 'https://www.airdna.co',
        'sec-fetch-site': 'same-site',
        'sec-fetch-mode': 'cors',
        'sec-fetch-dest': 'empty',
        'referer': 'https://www.airdna.co/',
        'accept-language': 'en-US,en;q=0.9',
    }
    search = ''
    if cityId != '':
        search = ('city_id', str(cityId))
    if regionId != '':
        search = ('region_id', str(regionId))
    params = (
        ('access_token', 'NzI0NDc2|7aef649379de442792bb19a20eee01a5'),
        search,
        ('start_month', '10'),
        ('start_year', '2018'),
        ('number_of_months', '36'),
        ('currency', 'native'),
        ('show_regions', 'true'),
    )

    res = requests.get('https://api.airdna.co/v1/market/property_list', headers=headers, params=params)
    res = res.json()
    if cityId != '':
        areaName = res['area_info']['geom']['name']['city'].replace('/', '_')
    if regionId != '':
        areaName = res['area_info']['geom']['name']['region'].replace('/', '_')
    print(areaName)
    return res['properties'], areaName
# {'airbnb_property_id': 7564825, 'rating': 9.5, 'property_type': 'Rental unit', 'bathrooms': 1.0, 'title': 'Logan
# Square 3 bedrms with parking', 'occ': 'permission_denied', 'bedrooms': 3, 'accommodates': 6, 'adr': 194.71153,
# 'img_cover': 'https://a0.muscache.com/im/pictures/95994001/4e7e4f88_original.jpg?aki_policy=x_large', 'reviews':
# 112, 'revenue': 'permission_denied', 'homeaway_property_id': None, 'longitude': -87.72377, 'regions': {
# 'neighborhood_ids': [126665], 'zipcode_ids': [23434]}, 'latitude': 41.92484, 'm_homeaway_property_id': None,
# 'days_available': 159, 'id': 21075, 'room_type': 'Entire home/apt', 'platforms': {'airbnb_property_id': 7564825,
# 'homeaway_property_id': None}}


counter = {0: 0,
           1: 0,
           2: 0,
           3: 0,
           4: 0}
AirdnaRes, areaName = Airdna(regionId=126703) # Set the city or region ID here
workbook = load_workbook(filename="../Template.xlsx")
sheet = workbook['Template']
sheet.title = areaName + " PY"
min_days = 250
min_occ = .50
min_reviews = 10
for listing in AirdnaRes:

    if listing["days_available"] < min_days or listing['room_type'] \
            == 'Private room' or listing['room_type'] == 'Shared room"' or listing['reviews'] <= min_reviews or listing['occ'] < min_occ:
        continue

    if listing["bedrooms"] >= len(counter):
        counter[len(counter)-1] += 1
        count = counter[len(counter)-1]
    else:
        counter[listing["bedrooms"]] += 1
        count = counter[listing["bedrooms"]]

    if listing["bedrooms"] == 0:
        sheet.cell(row=2, column=4 + count).value = "Listing " + str(count)
        sheet.cell(row=2, column=4 + count).comment = Comment(
            "\n\nAirDna:" + json.dumps(listing, indent=4), 'PY',
            948, 260)
        sheet.cell(row=2, column=4 + count).hyperlink = "https://www.airbnb.com/rooms/" + str(
            listing["airbnb_property_id"])
        sheet.cell(row=3, column=4 + count).value = listing["occ"]
        sheet.cell(row=4, column=4 + count).value = listing["adr"]

    elif listing["bedrooms"] == 1:
        sheet.cell(row=8, column=4 + count).value = "Listing " + str(count)
        sheet.cell(row=8, column=4 + count).comment = Comment(
             "\n\nAirDna:" + json.dumps(listing, indent=4), 'PY',
            948, 260)
        sheet.cell(row=8, column=4 + count).hyperlink = "https://www.airbnb.com/rooms/" + str(
            listing["airbnb_property_id"])
        sheet.cell(row=9, column=4 + count).value = listing["occ"]
        sheet.cell(row=10, column=4 + count).value = listing["adr"]

    elif listing["bedrooms"] == 2:
        sheet.cell(row=14, column=4 + count).value = "Listing " + str(count)
        sheet.cell(row=14, column=4 + count).comment = Comment(
            "\n\nAirDna:" + json.dumps(listing, indent=4), 'PY',
            948, 260)
        sheet.cell(row=14, column=4 + count).hyperlink = "https://www.airbnb.com/rooms/" + str(
            listing["airbnb_property_id"])
        sheet.cell(row=15, column=4 + count).value = listing["occ"]
        sheet.cell(row=16, column=4 + count).value = listing["adr"]

    elif listing["bedrooms"] == 3:
        sheet.cell(row=20, column=4 + count).value = "Listing " + str(count)
        sheet.cell(row=20, column=4 + count).comment = Comment(
             "\n\nAirDna:" + json.dumps(listing, indent=4), 'PY',
            948, 260)
        sheet.cell(row=20, column=4 + count).hyperlink = "https://www.airbnb.com/rooms/" + str(
            listing["airbnb_property_id"])
        sheet.cell(row=21, column=4 + count).value = listing["occ"]
        sheet.cell(row=22, column=4 + count).value = listing["adr"]

    elif listing["bedrooms"] >= 4:
        sheet.cell(row=26, column=11 + count).value = "Listing " + str(count)
        sheet.cell(row=26, column=11 + count).comment = Comment(
            "\n\nAirDna:" + json.dumps(listing, indent=4), 'PY',
            948, 260)
        sheet.cell(row=26, column=11 + count).hyperlink = "https://www.airbnb.com/rooms/" + str(
            listing["airbnb_property_id"])
        sheet.cell(row=27, column=11 + count).value = listing["occ"]
        sheet.cell(row=28, column=11 + count).value = listing["adr"]

    else:
        print("Bedroom>3 ----> " + json.dumps(listing, indent=4))

workbook.save('../'+areaName+"_Airdna_"+str(min_days)+'days.xlsx')
