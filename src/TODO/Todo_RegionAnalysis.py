from openpyxl import load_workbook
import requests
from openpyxl.comments import Comment
import json
from cityList import cityMain


def all_the_rooms(room_id):
    all_headers = {
        'authority': 'analytics.alltherooms.com',
        'accept': '*/*',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Safari/537.36',
        'content-type': 'application/json',
        'sec-gpc': '1',
        'origin': 'https://analytics.alltherooms.com',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-mode': 'cors',
        'sec-fetch-dest': 'empty',
        'referer': 'https://analytics.alltherooms.com/market/1040092',
        'accept-language': 'en-US,en;q=0.9',
        'cookie': '_fbp=fb.1.1631210716254.41325750; __stripe_mid=7244b495-1bc1-4083-8fd5-9ef57511a50cf85d79; future_data_feature=showed; account_v2=set; sign_in_popup_experiment=new; splitiokey=6999ca39-3ca4-476a-97b4-a2125143108b; sessionId=1633552823969; __stripe_sid=a3e30eb9-b2c9-46e3-8f36-e31d4d3d6fb6a9c6fd; atrsession=86739fea0cc09c9fb2c3bdc2a1d445c02e0b56fd6136ee13860585539e02f2ed8cdc43717ab8a35dc71da45fe884340da8060cae50405c8ad5ced3777282ca4ba9f90304eab03cfbc298ab11006e347d97a5154d036ab7c7babe91561f8654553cc9676d551ca1f71ded2f7787ef62c94177e628dcdfa174ef236d5f0177891a11f7882011cbdd31a58921d0790619; AWSALB=CiwgRT3i3FvtUY6IlovWNPy9MbxsSlvHBFTbWmU8afc4Gq2LIxHiuXNkWpSNuVCJbc62c4eReFetD+nx63G67K4oP8cm6XodeZObi1kuQHip6ZJOMB40/UCeEU5S; AWSALBCORS=CiwgRT3i3FvtUY6IlovWNPy9MbxsSlvHBFTbWmU8afc4Gq2LIxHiuXNkWpSNuVCJbc62c4eReFetD+nx63G67K4oP8cm6XodeZObi1kuQHip6ZJOMB40/UCeEU5S; lastSentEventId=1633553610205',
    }

    data = '{"operationName":"getListingDetails","variables":{"uid":"' + str(
        room_id) + '","providerId":"airbnb","month":"2021-08-01"},"query":"query getListingDetails(' \
                   '$areaId: Int, $uid: String!, $providerId: String!, $trackedUid: String, $trackedProviderId: ' \
                   'String, $month: String!) {\\n  listing(areaId: $areaId, uid: $uid, providerId: $providerId, ' \
                   'trackedUid: $trackedUid, trackedProviderId: $trackedProviderId, month: $month) {\\n    name\\n    ' \
                   'rating\\n    areaName\\n    areaId\\n    uid\\n    providerId\\n    name\\n    arrangementType\\n ' \
                   '   instantBook\\n    isManaged\\n    latitude\\n    longitude\\n    url\\n    sleeps\\n    ' \
                   'bedrooms\\n    bathrooms\\n    image {\\n      t\\n      n\\n      __typename\\n    }\\n    vrps ' \
                   '{\\n      value\\n      month\\n      __typename\\n    }\\n    isSuperhost\\n    dailyRate\\n    ' \
                   'occupancyRate\\n    trackedId\\n    rating\\n    reviewsCount\\n    beds\\n    images {\\n      ' \
                   'n\\n      caption\\n      __typename\\n    }\\n    hostName\\n    description\\n    amenities\\n  ' \
                   '  childrenAllowed\\n    eventsAllowed\\n    smokingAllowed\\n    petsAllowed\\n    checkInTime\\n ' \
                   '   checkOutTime\\n    cleaningFee\\n    weeklyDiscountFactor\\n    monthlyDiscountFactor\\n    ' \
                   'scores {\\n      areaId\\n      score\\n      difference\\n      description\\n      areaType\\n  ' \
                   '    __typename\\n    }\\n    vrpsHistory {\\n      areaId\\n      areaType\\n      scores {\\n    ' \
                   '    value\\n        month\\n        __typename\\n      }\\n      __typename\\n    }\\n    ' \
                   '__typename\\n  }\\n}\\n"} '

    all_response = requests.post('https://analytics.alltherooms.com/graphql', headers=all_headers, data=data)
    return all_response.json()["data"]["listing"]
    # {'name': 'Best in Chicago, private, amazing garden space', 'rating': 100, 'areaName': 'Logan Square, Chicago,
    # IL 60647, United States', 'areaId': 1040092, 'uid': '189821', 'providerId': 'airbnb', 'arrangementType': 'Entire
    # Home', 'instantBook': None, 'isManaged': None, 'latitude': 41.92918, 'longitude': -87.70219,
    # 'url': 'https://www.airbnb.com/rooms/189821', 'sleeps': 5, 'bedrooms': 2, 'bathrooms': 1, 'image': {'t': None,
    # 'n': 'https://a0.muscache.com/im/pictures/e6324b08-d6c8-4540-8164-410a96eadd2b.jpg', '__typename': 'Image'},
    # 'vrps': {'value': 188, 'month': None, '__typename': 'VrpsScore'}, 'isSuperhost': True, 'dailyRate': 191.814516129,
    # 'occupancyRate': 0.373494, 'trackedId': None, 'reviewsCount': 551, 'beds': 3, ...................


def Airdna(cityId='', regionId=''):
    headers = {
        'authority': 'api.airdna.co',
        'sec-ch-ua': '"Chromium";v="94", "Google Chrome";v="94", ";Not A Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/94.0.4606.71 Safari/537.36',
        'sec-ch-ua-platform': '"Windows"',
        'accept': '*/*',
        'origin': 'https://www.airdna.co',
        'sec-fetch-site': 'same-site',
        'sec-fetch-mode': 'cors',
        'sec-fetch-dest': 'empty',
        'referer': 'https://www.airdna.co/',
        'accept-language': 'en-US,en;q=0.9',
    }
    search = ''
    if cityId != '':
        search = ('city_id', str(cityId))
    if regionId != '':
        search = ('region_id', str(regionId))
    params = (
        ('access_token', 'MjkxMTI|8b0178bf0e564cbf96fc75b8518a5375'),
        search,
        ('start_month', '10'),
        ('start_year', '2018'),
        ('number_of_months', '36'),
        ('currency', 'native'),
        ('show_regions', 'true'),
    )

    res = requests.get('https://api.airdna.co/v1/market/property_list', headers=headers, params=params)
    res = res.json()
    if cityId != '':
        areaName = res['area_info']['geom']['name']['city'].replace('/', '_')
    if regionId != '':
        areaName = res['area_info']['geom']['name']['region'].replace('/', '_')
    print(areaName)
    return res['properties'], areaName
# {'airbnb_property_id': 7564825, 'rating': 9.5, 'property_type': 'Rental unit', 'bathrooms': 1.0, 'title': 'Logan
# Square 3 bedrms with parking', 'occ': 'permission_denied', 'bedrooms': 3, 'accommodates': 6, 'adr': 194.71153,
# 'img_cover': 'https://a0.muscache.com/im/pictures/95994001/4e7e4f88_original.jpg?aki_policy=x_large', 'reviews':
# 112, 'revenue': 'permission_denied', 'homeaway_property_id': None, 'longitude': -87.72377, 'regions': {
# 'neighborhood_ids': [126665], 'zipcode_ids': [23434]}, 'latitude': 41.92484, 'm_homeaway_property_id': None,
# 'days_available': 159, 'id': 21075, 'room_type': 'Entire home/apt', 'platforms': {'airbnb_property_id': 7564825,
# 'homeaway_property_id': None}}


regions = cityMain()
for region in regions:
    for city in regions[region][1]:

        areaName = ''
        counter = {0: 0,
                   1: 0,
                   2: 0,
                   3: 0}
        AirdnaRes, areaName = Airdna(regionId=regions[region][1][city])
        workbook = load_workbook(filename="./Template.xlsx")
        sheet = workbook['Template']
        sheet.title = areaName + " PY"
        for listing in AirdnaRes:

            allTheRoomsRes = all_the_rooms(listing['airbnb_property_id'])
            if listing["days_available"] < 275 or listing["bedrooms"] >= len(counter) or listing['room_type'] \
                    == 'Private room' or listing['room_type'] == 'Shared room"' or allTheRoomsRes is None:
                continue

            allTheRoomsRes.pop('images');
            allTheRoomsRes.pop('amenities');
            allTheRoomsRes.pop('description');
            allTheRoomsRes.pop('vrpsHistory');
            allTheRoomsRes.pop('scores')
            # areaName = allTheRoomsRes['areaName'].split(',')[0]
            counter[listing["bedrooms"]] += 1
            count = counter[listing["bedrooms"]]

            if listing["bedrooms"] == 0:
                sheet.cell(row=2, column=4 + count).value = "Listing " + str(count)
                sheet.cell(row=2, column=4 + count).comment = Comment(
                    "AllTheRooms:" + json.dumps(allTheRoomsRes, indent=4) + "\n\nAirDna:" + json.dumps(listing, indent=4), 'PY',
                    948, 260)
                sheet.cell(row=2, column=4 + count).hyperlink = "https://www.airbnb.com/rooms/" + str(
                    listing["airbnb_property_id"])
                sheet.cell(row=3, column=4 + count).value = allTheRoomsRes["occupancyRate"]
                sheet.cell(row=4, column=4 + count).value = allTheRoomsRes["dailyRate"]

            elif listing["bedrooms"] == 1:
                sheet.cell(row=8, column=4 + count).value = "Listing " + str(count)
                sheet.cell(row=8, column=4 + count).comment = Comment(
                    "AllTheRooms:" + json.dumps(allTheRoomsRes, indent=4) + "\n\nAirDna:" + json.dumps(listing, indent=4), 'PY',
                    948, 260)
                sheet.cell(row=8, column=4 + count).hyperlink = "https://www.airbnb.com/rooms/" + str(
                    listing["airbnb_property_id"])
                sheet.cell(row=9, column=4 + count).value = allTheRoomsRes["occupancyRate"]
                sheet.cell(row=10, column=4 + count).value = allTheRoomsRes["dailyRate"]

            elif listing["bedrooms"] == 2:
                sheet.cell(row=14, column=4 + count).value = "Listing " + str(count)
                sheet.cell(row=14, column=4 + count).comment = Comment(
                    "AllTheRooms:" + json.dumps(allTheRoomsRes, indent=4) + "\n\nAirDna:" + json.dumps(listing, indent=4), 'PY',
                    948, 260)
                sheet.cell(row=14, column=4 + count).hyperlink = "https://www.airbnb.com/rooms/" + str(
                    listing["airbnb_property_id"])
                sheet.cell(row=15, column=4 + count).value = allTheRoomsRes["occupancyRate"]
                sheet.cell(row=16, column=4 + count).value = allTheRoomsRes["dailyRate"]

            elif listing["bedrooms"] == 3:
                sheet.cell(row=20, column=4 + count).value = "Listing " + str(count)
                sheet.cell(row=20, column=4 + count).comment = Comment(
                    "AllTheRooms:" + json.dumps(allTheRoomsRes, indent=4) + "\n\nAirDna:" + json.dumps(listing, indent=4), 'PY',
                    948, 260)
                sheet.cell(row=20, column=4 + count).hyperlink = "https://www.airbnb.com/rooms/" + str(
                    listing["airbnb_property_id"])
                sheet.cell(row=21, column=4 + count).value = allTheRoomsRes["occupancyRate"]
                sheet.cell(row=22, column=4 + count).value = allTheRoomsRes["dailyRate"]

            else:
                print("Bedroom>3 ----> " + json.dumps(listing, indent=4))

        workbook.save('../'+areaName+'.xlsx')
